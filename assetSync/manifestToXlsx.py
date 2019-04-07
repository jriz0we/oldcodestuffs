import os
import json
from openpyxl import load_workbook

XLSXPATH  = 'card-heroes-excel/game/card-heroes-audiovisuals.xlsx'
MANIFEST_PATH = 'card-heroes-assetpipe/unityproject/Assets/exported_assets/manifests/merged/data_manifest.json'
manifestpath = os.path.join(os.environ['HOME'],MANIFEST_PATH)
workbookpath = os.path.join(os.environ['HOME'],XLSXPATH)

vfxassets = []
cardassets = []
chibiassets = []
sfxassets = []
iconassets = []
worksheetTargets = ['cardprefabs','chibisprefabs','vfxprefabs','sfxprefabs','iconprefabs']

def readSheet(worksheet):
    print('reading existing sheet:',worksheet.title)
    existingList = []
    for row in worksheet.iter_rows(min_row = 1, max_col=3):
        for cell in row:
            existingList.append(cell.value)
    return existingList

def diffLists(sourceList, newList):
    if set(sourceList) != set(newList):
        diffs = set(newList).difference(sourceList)
        return list(diffs)
    else:
        return False

def fillSheetFromList(assetList, worksheet):
    print('writing out sheet ',worksheet)
    col = 1
    row = 1
    for asset in assetList:
        worksheet.cell(column=col, row = row, value = asset)
        row += 1
    return worksheet


def updateSheets(assetList,sheetName):
    existingList = readSheet(sheetName)
    diffs = diffLists(existingList,assetList)
    if diffs:
        print('assets to add:')
        for asset in sorted(diffs):
            print(asset)
        fillSheetFromList(assetList,sheetName)


print('Reading manifest...')
with open(manifestpath) as json_file:
    data = json.load(json_file)
    for p in data['Assets']:
        #exclude data manifest objects
        if 'data_' not in p['AssetName'] and '(1)' not in p['AssetName']:
            if p['AssetName'].startswith('card_'):cardassets.append(p['AssetName'])
            if p['AssetName'].startswith('store_'):iconassets.append(p['AssetName'])
            if p['AssetName'].startswith('narration_'):iconassets.append(p['AssetName'])
            if p['AssetName'].startswith('reward_'):iconassets.append(p['AssetName'])
            if p['AssetName'].startswith('social_'):iconassets.append(p['AssetName'])
            if p['AssetName'].startswith('vfx_'):vfxassets.append(p['AssetName'])
            if p['AssetName'].startswith('sfx_'):sfxassets.append(p['AssetName'])
            if p['AssetName'].startswith('mus_'):print('hey music!')
            #add everything not data to chibiassets since it doesnt have a prefix
            else: chibiassets.append(p['AssetName'])

#sort the lists
iconassets=sorted(set(iconassets))
cardassets= sorted(set(cardassets))
vfxassets=sorted(set(vfxassets))
sfxassets=sorted(set(sfxassets))

#remove the other 3 lists form the chibi list
chibiassets = sorted(list (set(chibiassets) - set(vfxassets) -  set(cardassets) - set(sfxassets)-set(iconassets)))

assetLists = [cardassets,chibiassets,vfxassets,sfxassets,iconassets]
wb = load_workbook(filename = workbookpath)

for assettype,sheet in zip(assetLists,worksheetTargets):
    updateSheets(assettype,wb[sheet])

print('saving file:', workbookpath)
wb.save(filename = workbookpath)

